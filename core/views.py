import io
import json
from datetime import datetime, time as dt_time
from functools import wraps
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login as auth_login, logout as auth_logout, authenticate
from django.contrib.auth.decorators import login_required
from django.contrib.auth.forms import AuthenticationForm
from django.http import HttpResponse, HttpResponseForbidden, JsonResponse
from django.utils import timezone
from django.contrib import messages
from django.db import transaction

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .models import User, Machine, ChecklistSession, ChecklistTimelineLog, ChecklistItemValue
from .forms import RegisterForm, MachineForm, ChecklistStartForm
from .checklist_items import CHECKLIST_ITEMS

# =====================================================================
# AUTHENTICATION VIEWS
# =====================================================================

def login_view(request):
    if request.user.is_authenticated:
        return redirect('dashboard')
    
    if request.method == 'POST':
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            user = form.get_user()
            if user is not None:
                if user.is_active:
                    auth_login(request, user)
                    return redirect('dashboard')
        else:
            # Check if user exists and is inactive with correct credentials
            username = request.POST.get('username')
            password = request.POST.get('password')
            user = User.objects.filter(username=username).first()
            if user and not user.is_active:
                if user.check_password(password):
                    messages.warning(
                        request,
                        "Seu cadastro foi realizado com sucesso, mas aguarda a aprovação de um Administrador/Diretor para liberar o acesso."
                    )
                    return render(request, 'core/login.html', {'form': form})
            
            messages.error(request, "Por favor, insira um usuário e senha corretos.")
    else:
        form = AuthenticationForm()
    return render(request, 'core/login.html', {'form': form})

def register_view(request):
    if request.user.is_authenticated:
        return redirect('dashboard')
    if request.method == 'POST':
        form = RegisterForm(request.POST)
        if form.is_valid():
            user = form.save()
            messages.success(
                request,
                "Seu cadastro foi realizado com sucesso, mas aguarda a aprovação de um Administrador/Diretor para liberar o acesso."
            )
            return redirect('login')
    else:
        form = RegisterForm()
    return render(request, 'core/register.html', {'form': form})

# =====================================================================
# MACHINE CRUD VIEWS
# =====================================================================

def machine_management_required(view_func):
    @wraps(view_func)
    def _wrapped_view(request, *args, **kwargs):
        if request.user.is_authenticated and (request.user.is_superuser or request.user.specialty in ['Diretor', 'Analista']):
            return view_func(request, *args, **kwargs)
        messages.error(request, "Você não tem permissão para acessar esta página.")
        return redirect('dashboard')
    return _wrapped_view

@login_required
@machine_management_required
def machine_list(request):
    machines = Machine.objects.all().order_by('name')
    return render(request, 'core/machine_list.html', {'machines': machines})

@login_required
@machine_management_required
def machine_create(request):
    if request.method == 'POST':
        form = MachineForm(request.POST)
        if form.is_valid():
            machine = form.save()
            messages.success(request, f"Máquina '{machine.name}' cadastrada com sucesso!")
            return redirect('machine_list')
    else:
        form = MachineForm()
    return render(request, 'core/machine_form.html', {'form': form, 'title': 'Cadastrar Máquina'})

@login_required
@machine_management_required
def machine_update(request, pk):
    machine = get_object_or_404(Machine, pk=pk)
    if request.method == 'POST':
        form = MachineForm(request.POST, instance=machine)
        if form.is_valid():
            form.save()
            messages.success(request, f"Máquina '{machine.name}' atualizada com sucesso!")
            return redirect('machine_list')
    else:
        form = MachineForm(instance=machine)
    return render(request, 'core/machine_form.html', {'form': form, 'title': 'Editar Máquina', 'machine': machine})

@login_required
@machine_management_required
def machine_delete(request, pk):
    machine = get_object_or_404(Machine, pk=pk)
    if request.method == 'POST':
        machine.delete()
        messages.success(request, f"Máquina '{machine.name}' excluída com sucesso!")
        return redirect('machine_list')
    return render(request, 'core/machine_confirm_delete.html', {'machine': machine})

# =====================================================================
# DASHBOARD VIEW
# =====================================================================

@login_required
def dashboard(request):
    user = request.user

    # --- FILTRO POR PERÍODO (parâmetros GET) ---
    today = timezone.localdate()
    default_inicio = today.replace(day=1)  # 1º dia do mês corrente
    default_fim = today

    data_inicio_str = request.GET.get('data_inicio', '')
    data_fim_str = request.GET.get('data_fim', '')

    try:
        data_inicio = datetime.strptime(data_inicio_str, '%Y-%m-%d').date() if data_inicio_str else default_inicio
    except ValueError:
        data_inicio = default_inicio

    try:
        data_fim = datetime.strptime(data_fim_str, '%Y-%m-%d').date() if data_fim_str else default_fim
    except ValueError:
        data_fim = default_fim

    # Converter para datetime aware para comparação com DateTimeField
    dt_inicio = timezone.make_aware(datetime.combine(data_inicio, dt_time.min))
    dt_fim = timezone.make_aware(datetime.combine(data_fim, dt_time.max))

    # --- FILTROS AVANÇADOS (parâmetros GET) ---
    maquina_id = request.GET.get('maquina', '')
    inspector_id = request.GET.get('inspector', '')
    leader_id = request.GET.get('leader_filter', '')
    status_filter = request.GET.get('status_filter', '')
    nc_query = request.GET.get('nc_query', '')

    # --- QUERYSET DE SESSÕES (filtrada por período) ---
    if user.is_technician:
        # Technicians only see checklists they inspected
        sessions = ChecklistSession.objects.filter(
            inspector=user, created_at__gte=dt_inicio, created_at__lte=dt_fim
        ).select_related('machine', 'inspector', 'leader').order_by('-created_at')
    else:
        # Leaders, Analysts, and Directors see all checklists
        sessions = ChecklistSession.objects.filter(
            created_at__gte=dt_inicio, created_at__lte=dt_fim
        ).select_related('machine', 'inspector', 'leader').order_by('-created_at')

    # Aplicar filtros adicionais combinados
    if maquina_id:
        sessions = sessions.filter(machine_id=maquina_id)
    if inspector_id:
        sessions = sessions.filter(inspector_id=inspector_id)
    if leader_id:
        sessions = sessions.filter(leader_id=leader_id)
    if status_filter:
        sessions = sessions.filter(status=status_filter)
    if nc_query:
        sessions = sessions.filter(items__status='NC', items__item_name__icontains=nc_query).distinct()

    # Dropdowns de Filtro
    machines = Machine.objects.all().order_by('name')
    inspectors = User.objects.filter(is_active=True).exclude(specialty='Lider').order_by('first_name', 'username')
    leaders = User.objects.filter(specialty='Lider', is_active=True).order_by('first_name', 'username')

    # --- KPIs CALCULADOS NO BACKEND ---
    approved_count = sessions.filter(status='APROVADO').count()
    active_count = sessions.filter(status__in=['IN_PROGRESS', 'PAUSED']).count()
    total_ncs = 0

    # Prepare non-conformity list for each session
    for s in sessions:
        nc_items = s.items.filter(status='NC')
        s.nc_list = [item.item_name for item in nc_items]
        s.nc_summary = ", ".join(s.nc_list) if s.nc_list else "Nenhuma"
        total_ncs += len(s.nc_list)

    # --- FILAS DE APROVAÇÃO (NÃO filtradas por período — operacional) ---
    # Get pending checklists for Lider approval
    pending_leader = []
    if user.is_leader or user.is_superuser:
        pending_leader = ChecklistSession.objects.filter(status='AGUARDANDO_LIDER').select_related('machine', 'inspector').order_by('-created_at')
        for s in pending_leader:
            nc_items = s.items.filter(status='NC')
            s.nc_list = [item.item_name for item in nc_items]
            s.nc_summary = ", ".join(s.nc_list) if s.nc_list else "Nenhuma"

    # Get checklists rejected by Leader for Analyst review
    pending_analyst = []
    if user.is_analyst or user.is_director or user.is_superuser:
        pending_analyst = ChecklistSession.objects.filter(status='REPROVADO_LIDER').select_related('machine', 'inspector').order_by('-created_at')
        for s in pending_analyst:
            nc_items = s.items.filter(status='NC')
            s.nc_list = [item.item_name for item in nc_items]
            s.nc_summary = ", ".join(s.nc_list) if s.nc_list else "Nenhuma"

    # Get checklists rejected by Analyst for Technician corrections
    pending_corrections = []
    if user.is_technician:
        pending_corrections = ChecklistSession.objects.filter(status='REPROVAR_FINAL_REFAZER', inspector=user).select_related('machine', 'inspector').order_by('-created_at')
        for s in pending_corrections:
            nc_items = s.items.filter(status='NC')
            s.nc_list = [item.item_name for item in nc_items]
            s.nc_summary = ", ".join(s.nc_list) if s.nc_list else "Nenhuma"

    return render(request, 'core/dashboard.html', {
        'sessions': sessions,
        'approved_count': approved_count,
        'active_count': active_count,
        'total_ncs': total_ncs,
        'data_inicio': data_inicio.isoformat(),
        'data_fim': data_fim.isoformat(),
        'pending_leader': pending_leader,
        'pending_analyst': pending_analyst,
        'pending_corrections': pending_corrections,
        'machines': machines,
        'inspectors': inspectors,
        'leaders': leaders,
        'maquina_id': maquina_id,
        'inspector_id': inspector_id,
        'leader_id': leader_id,
        'status_filter': status_filter,
        'nc_query': nc_query,
    })


@login_required
@machine_management_required
def analytical_dashboard(request):
    today = timezone.localdate()
    default_inicio = today.replace(day=1)  # 1º dia do mês corrente
    default_fim = today

    data_inicio_str = request.GET.get('data_inicio', '')
    data_fim_str = request.GET.get('data_fim', '')

    try:
        data_inicio = datetime.strptime(data_inicio_str, '%Y-%m-%d').date() if data_inicio_str else default_inicio
    except ValueError:
        data_inicio = default_inicio

    try:
        data_fim = datetime.strptime(data_fim_str, '%Y-%m-%d').date() if data_fim_str else default_fim
    except ValueError:
        data_fim = default_fim

    # Converter para datetime aware para comparação com DateTimeField
    dt_inicio = timezone.make_aware(datetime.combine(data_inicio, dt_time.min))
    dt_fim = timezone.make_aware(datetime.combine(data_fim, dt_time.max))

    # Query de checklists no período
    sessions = ChecklistSession.objects.filter(
        created_at__gte=dt_inicio, created_at__lte=dt_fim
    ).select_related('machine')

    # Filter completed sessions
    completed_sessions = [s for s in sessions if s.completed_at]

    # --- 001. Gráfico de Evolução dos Checklists Realizados ---
    delta_days = (data_fim - data_inicio).days
    evolution_labels = []
    evolution_data = []

    if delta_days <= 30:
        # Group by day
        from datetime import timedelta
        date_list = [data_inicio + timedelta(days=x) for x in range(delta_days + 1)]
        counts = {d: 0 for d in date_list}
        for s in completed_sessions:
            s_date = timezone.localdate(s.completed_at)
            if s_date in counts:
                counts[s_date] += 1
        evolution_labels = [d.strftime('%d/%m') for d in date_list]
        evolution_data = [counts[d] for d in date_list]
    elif delta_days <= 120:
        # Group by week
        from datetime import timedelta
        start_monday = data_inicio - timedelta(days=data_inicio.weekday())
        week_list = []
        curr = start_monday
        while curr <= data_fim:
            week_list.append(curr)
            curr += timedelta(days=7)

        counts = {w: 0 for w in week_list}
        for s in completed_sessions:
            s_date = timezone.localdate(s.completed_at)
            s_monday = s_date - timedelta(days=s_date.weekday())
            if s_monday in counts:
                counts[s_monday] += 1
            elif s_monday < start_monday and s_date >= data_inicio:
                counts[start_monday] += 1

        evolution_labels = [f"Sem {w.strftime('%d/%m')}" for w in week_list]
        evolution_data = [counts[w] for w in week_list]
    else:
        # Group by month
        curr_year = data_inicio.year
        curr_month = data_inicio.month
        month_list = []
        while (curr_year < data_fim.year) or (curr_year == data_fim.year and curr_month <= data_fim.month):
            month_list.append((curr_year, curr_month))
            if curr_month == 12:
                curr_month = 1
                curr_year += 1
            else:
                curr_month += 1

        counts = {m: 0 for m in month_list}
        month_names = {
            1: 'Jan', 2: 'Fev', 3: 'Mar', 4: 'Abr', 5: 'Mai', 6: 'Jun',
            7: 'Jul', 8: 'Ago', 9: 'Set', 10: 'Out', 11: 'Nov', 12: 'Dez'
        }
        for s in completed_sessions:
            s_date = timezone.localdate(s.completed_at)
            key = (s_date.year, s_date.month)
            if key in counts:
                counts[key] += 1
        evolution_labels = [f"{month_names[m]}/{y}" for (y, m) in month_list]
        evolution_data = [counts[m] for m in month_list]

    # --- 002. Gráfico de Quantidade de Problemas Encontrados por Máquina ---
    machines = Machine.objects.all().order_by('name')
    problems_by_machine = {m.name: 0 for m in machines}

    nc_items = ChecklistItemValue.objects.filter(
        session__in=sessions,
        status='NC'
    ).select_related('session__machine')

    for item in nc_items:
        machine_name = item.session.machine.name
        problems_by_machine[machine_name] = problems_by_machine.get(machine_name, 0) + 1

    sorted_problems = sorted(problems_by_machine.items(), key=lambda x: x[1], reverse=True)
    machine_labels = [x[0] for x in sorted_problems]
    machine_data = [x[1] for x in sorted_problems]

    # --- 003. Gráfico de Distribuição dos Tipos de Falha ---
    section_map = {
        'HIDRAULICA': 'Hidráulica',
        'PNEUMATICA': 'Pneumática',
        'SISTEMA_VULCANIZACAO': 'Vulcanização',
        'SISTEMA_VAPOR': 'Vapor',
        'ESTRUTURA': 'Estrutura',
        'ELETRICA': 'Elétrica',
    }

    problems_by_section = {section_map[k]: 0 for k in section_map}
    for item in nc_items:
        clean_section = section_map.get(item.section, item.section)
        problems_by_section[clean_section] = problems_by_section.get(clean_section, 0) + 1

    section_labels = list(problems_by_section.keys())
    section_data = list(problems_by_section.values())

    # --- 004. Gráfico de Reincidência de Problemas ---
    curr_year = data_inicio.year
    curr_month = data_inicio.month
    months_list = []
    while (curr_year < data_fim.year) or (curr_year == data_fim.year and curr_month <= data_fim.month):
        months_list.append((curr_year, curr_month))
        if curr_month == 12:
            curr_month = 1
            curr_year += 1
        else:
            curr_month += 1

    month_names = {
        1: 'Jan', 2: 'Fev', 3: 'Mar', 4: 'Abr', 5: 'Mai', 6: 'Jun',
        7: 'Jul', 8: 'Ago', 9: 'Set', 10: 'Out', 11: 'Nov', 12: 'Dez'
    }
    reincidence_labels = [f"{month_names[m]}/{y}" for (y, m) in months_list]

    section_datasets = {
        sec_code: {
            'label': sec_name,
            'data': [0] * len(months_list)
        } for sec_code, sec_name in section_map.items()
    }

    for item in nc_items:
        if item.session.completed_at:
            s_date = timezone.localdate(item.session.completed_at)
            key = (s_date.year, s_date.month)
            if key in months_list:
                month_idx = months_list.index(key)
                if item.section in section_datasets:
                    section_datasets[item.section]['data'][month_idx] += 1

    reincidence_datasets = list(section_datasets.values())

    return render(request, 'core/analytical_dashboard.html', {
        'data_inicio': data_inicio.isoformat(),
        'data_fim': data_fim.isoformat(),
        'evolution_labels': json.dumps(evolution_labels),
        'evolution_data': json.dumps(evolution_data),
        'machine_labels': json.dumps(machine_labels),
        'machine_data': json.dumps(machine_data),
        'section_labels': json.dumps(section_labels),
        'section_data': json.dumps(section_data),
        'reincidence_labels': json.dumps(reincidence_labels),
        'reincidence_datasets': json.dumps(reincidence_datasets),
    })


# =====================================================================
# CHECKLIST EXECUTION VIEWS (State Machine Workflow)
# =====================================================================


@login_required
def checklist_start(request):
    if not request.user.can_create_checklist:
        return HttpResponseForbidden("Líderes de turno não possuem permissão para iniciar novos checklists.")

    # Active check: redirect if user has an in-progress checklist
    active_session = ChecklistSession.objects.filter(inspector=request.user, status='IN_PROGRESS').first()
    if active_session:
        messages.warning(request, "Você possui um checklist em andamento. Finalize ou pause antes de iniciar outro.")
        return redirect('checklist_execute', session_id=active_session.id)

    if request.method == 'POST':
        form = ChecklistStartForm(request.POST, user=request.user)
        if form.is_valid():
            with transaction.atomic():
                session = form.save(commit=False)
                session.inspector = request.user
                session.status = 'IN_PROGRESS'
                session.started_at = timezone.now()
                session.save()
                
                # Prepopulate/seed items
                for section_code, items in CHECKLIST_ITEMS.items():
                    for item in items:
                        ChecklistItemValue.objects.create(
                            session=session,
                            section=section_code,
                            item_name=item,
                            status=None,
                            observations=''
                        )
                
                # Write to audit timeline log
                ChecklistTimelineLog.objects.create(
                    session=session,
                    user=request.user,
                    action='START'
                )
            messages.success(request, "Inspeção iniciada!")
            return redirect('checklist_execute', session_id=session.id)
    else:
        form = ChecklistStartForm(user=request.user)
        
    context = {
        'form': form,
        'checklist_items_preview': CHECKLIST_ITEMS,
        'inspector': request.user,
    }
    return render(request, 'core/checklist_start.html', context)

@login_required
def checklist_execute(request, session_id):
    session = get_object_or_404(ChecklistSession, id=session_id)
    items = session.items.all().order_by('id')
    pause_logs = session.timeline_logs.filter(action='PAUSE').order_by('timestamp')
    
    sections_map = {
        'HIDRAULICA': 'Hidráulica (Mecânico)',
        'PNEUMATICA': 'Pneumática (Mecânico)',
        'SISTEMA_VULCANIZACAO': 'Sistema de Vulcanização (Mecânico)',
        'SISTEMA_VAPOR': 'Sistema de Vapor (Mecânico)',
        'ESTRUTURA': 'Estrutura (Mecânico)',
        'ELETRICA': 'Elétrica (Eletricista)',
    }
    
    # Check authorization to edit
    is_editable = session.status in ['IN_PROGRESS', 'PAUSED']
    if is_editable:
        if request.user != session.inspector and not (request.user.is_director or request.user.is_analyst or request.user.is_superuser):
            return HttpResponseForbidden("Você não tem permissão para editar o checklist de outro inspetor.")

    # Handle state transitions and execution POST actions
    if request.method == 'POST':
        action = request.POST.get('action')
        
        # Only allow modification if status is in_progress/paused
        if not is_editable:
            messages.error(request, "Este checklist não pode ser modificado no status atual.")
            return redirect('checklist_execute', session_id=session.id)
            
        if action == 'pause':
            pause_reason = request.POST.get('pause_reason', '').strip()
            general_obs = request.POST.get('general_observations', '').strip()
            if not pause_reason:
                messages.error(request, "É necessário informar o motivo da pausa.")
                return redirect('checklist_execute', session_id=session.id)
            
            with transaction.atomic():
                for item in items:
                    status_val = request.POST.get(f'status_{item.id}')
                    obs_val = request.POST.get(f'observations_{item.id}', '').strip()
                    item.status = status_val if status_val in ['C', 'NC'] else None
                    item.observations = obs_val
                    item.save()

                session.general_observations = general_obs
                session.status = 'PAUSED'
                session.paused_at = timezone.now()
                session.save()
                
                ChecklistTimelineLog.objects.create(
                    session=session,
                    user=request.user,
                    action='PAUSE',
                    pause_reason=pause_reason
                )
            messages.warning(request, "Checklist pausado. Suas alterações foram salvas.")
            return redirect('checklist_execute', session_id=session.id)
            
        elif action == 'continue':
            if session.status != 'PAUSED':
                messages.error(request, "Este checklist não está pausado.")
                return redirect('checklist_execute', session_id=session.id)
                
            with transaction.atomic():
                session.status = 'IN_PROGRESS'
                session.resumed_at = timezone.now()
                session.save()
                
                ChecklistTimelineLog.objects.create(
                    session=session,
                    user=request.user,
                    action='CONTINUE'
                )
            messages.success(request, "Checklist retomado!")
            return redirect('checklist_execute', session_id=session.id)
            
        elif action == 'finalize':
            errors = []
            updated_items = []
            general_obs = request.POST.get('general_observations', '').strip()
            
            with transaction.atomic():
                for item in items:
                    status_val = request.POST.get(f'status_{item.id}')
                    obs_val = request.POST.get(f'observations_{item.id}', '').strip()
                    
                    # Check validation: status is mandatory
                    if not status_val or status_val not in ['C', 'NC']:
                        errors.append(f"O item '{item.item_name}' ({sections_map.get(item.section)}) está pendente.")
                    # Check validation: observation is mandatory if NC
                    elif status_val == 'NC' and not obs_val:
                        errors.append(f"O item '{item.item_name}' ({sections_map.get(item.section)}) foi marcado como Não Conforme e exige uma observação justificando.")
                        
                    item.status = status_val if status_val in ['C', 'NC'] else None
                    item.observations = obs_val
                    updated_items.append(item)
                
                # Save progress even if validation failed
                for item in updated_items:
                    item.save()
                
                session.general_observations = general_obs
                session.save()
                    
                if errors:
                    # Regroup items for rendering with errors
                    grouped_items = {}
                    for section_code, section_name in sections_map.items():
                        grouped_items[section_code] = {
                            'name': section_name,
                            'items': items.filter(section=section_code)
                        }
                    
                    context = {
                        'session': session,
                        'grouped_items': grouped_items,
                        'errors': errors,
                        'pause_logs': pause_logs,
                    }
                    return render(request, 'core/checklist_form.html', context)
                
                # Transition status to AGUARDANDO_LIDER
                session.status = 'AGUARDANDO_LIDER'
                session.completed_at = timezone.now()
                session.save()
                
                ChecklistTimelineLog.objects.create(
                    session=session,
                    user=request.user,
                    action='FINISH'
                )
                
            messages.success(request, "Checklist finalizado! Aguardando aprovação do Líder de Turno.")
            return redirect('dashboard')

    # Group items for rendering
    grouped_items = {}
    for section_code, section_name in sections_map.items():
        grouped_items[section_code] = {
            'name': section_name,
            'items': items.filter(section=section_code)
        }
        
    context = {
        'session': session,
        'grouped_items': grouped_items,
        'pause_logs': pause_logs,
    }
    return render(request, 'core/checklist_form.html', context)

# =====================================================================
# APPROVAL WORKFLOW ACTION VIEWS
# =====================================================================

@login_required
def leader_decision(request, session_id):
    if not (request.user.is_leader or request.user.is_superuser):
        return HttpResponseForbidden("Apenas líderes podem aprovar/reprovar nesta etapa.")
    
    session = get_object_or_404(ChecklistSession, id=session_id)
    if session.status != 'AGUARDANDO_LIDER':
        messages.error(request, "Este checklist não está aguardando aprovação do líder.")
        return redirect('dashboard')
        
    if request.method == 'POST':
        decision = request.POST.get('decision')
        if decision == 'approve':
            session.status = 'APROVADO'
            session.save()
            ChecklistTimelineLog.objects.create(
                session=session,
                user=request.user,
                action='APPROVE_LIDER'
            )
            messages.success(request, f"Checklist #{session.id} aprovado com sucesso!")
        elif decision == 'reject':
            session.status = 'REPROVADO_LIDER'
            session.save()
            ChecklistTimelineLog.objects.create(
                session=session,
                user=request.user,
                action='REJECT_LIDER'
            )
            messages.warning(request, f"Checklist #{session.id} reprovado e encaminhado para revisão dos Analistas.")
        else:
            messages.error(request, "Decisão inválida.")
            
    return redirect('dashboard')

@login_required
def analyst_decision(request, session_id):
    if not (request.user.is_analyst or request.user.is_director or request.user.is_superuser):
        return HttpResponseForbidden("Apenas analistas ou diretores podem revisar checklists reprovados.")
    
    session = get_object_or_404(ChecklistSession, id=session_id)
    if session.status != 'REPROVADO_LIDER':
        messages.error(request, "Este checklist não está aguardando revisão do analista.")
        return redirect('dashboard')
        
    if request.method == 'POST':
        decision = request.POST.get('decision')
        if decision == 'approve':
            session.status = 'APROVADO'
            session.save()
            ChecklistTimelineLog.objects.create(
                session=session,
                user=request.user,
                action='APPROVE_ANALISTA'
            )
            messages.success(request, f"Checklist #{session.id} aprovado pelo Analista!")
        elif decision == 'reject':
            session.status = 'REPROVAR_FINAL_REFAZER'
            session.save()
            ChecklistTimelineLog.objects.create(
                session=session,
                user=request.user,
                action='REJECT_ANALISTA'
            )
            messages.warning(request, f"Checklist #{session.id} reprovado final e devolvido ao técnico para refazer.")
        else:
            messages.error(request, "Decisão inválida.")
            
    return redirect('dashboard')

@login_required
def checklist_reopen(request, session_id):
    session = get_object_or_404(ChecklistSession, id=session_id)
    
    # Only the original inspector or superuser/director/analyst can reopen
    if request.user != session.inspector and not (request.user.is_director or request.user.is_analyst or request.user.is_superuser):
        return HttpResponseForbidden("Você não tem permissão para reabrir este checklist.")
        
    if session.status != 'REPROVAR_FINAL_REFAZER':
        messages.error(request, "Este checklist não está no status que permite ser reaberto.")
        return redirect('dashboard')
        
    if request.method == 'POST':
        session.status = 'IN_PROGRESS'
        session.save()
        ChecklistTimelineLog.objects.create(
            session=session,
            user=request.user,
            action='REOPEN'
        )
        messages.success(request, f"Checklist #{session.id} reaberto. Você pode corrigir os itens agora.")
        return redirect('checklist_execute', session_id=session.id)
        
    return redirect('dashboard')

# =====================================================================
# USER APPROVAL VIEWS
# =====================================================================

@login_required
def user_approval_list(request):
    if not (request.user.is_director or request.user.is_analyst or request.user.is_superuser):
        return HttpResponseForbidden("Acesso exclusivo a Diretores e Analistas para aprovação de usuários.")
        
    pending_users = User.objects.filter(is_active=False).order_by('-date_joined')
    return render(request, 'core/user_approval_list.html', {'pending_users': pending_users})

@login_required
def approve_user(request, user_id):
    if not (request.user.is_director or request.user.is_analyst or request.user.is_superuser):
        return HttpResponseForbidden("Acesso exclusivo a Diretores e Analistas para ativação de usuários.")
        
    user_to_approve = get_object_or_404(User, id=user_id)
    if request.method == 'POST':
        user_to_approve.is_active = True
        user_to_approve.save()
        messages.success(request, f"Usuário {user_to_approve.username} ativado com sucesso!")
        
    return redirect('user_approval_list')

# =====================================================================
# EXCEL EXPORT VIEWS
# =====================================================================

@login_required
def export_checklist_excel(request, session_id):
    if not request.user.can_export_excel:
        return HttpResponseForbidden("Você não tem permissão para exportar checklists para Excel.")
        
    session = get_object_or_404(ChecklistSession, id=session_id)
    items = session.items.all().order_by('id')
    
    wb = Workbook()
    ws = wb.active
    ws.title = f"Checklist - ID {session.id}"
    
    # Formatting Fonts & Styles
    title_font = Font(name='Segoe UI', size=16, bold=True, color='FFFFFF')
    header_font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
    meta_label_font = Font(name='Segoe UI', size=10, bold=True, color='2c3e50')
    meta_value_font = Font(name='Segoe UI', size=10, color='333333')
    section_font = Font(name='Segoe UI', size=11, bold=True, color='2c3e50')
    item_font = Font(name='Segoe UI', size=10)
    
    # Fills
    title_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid') # Slate Blue
    header_fill = PatternFill(start_color='2F5597', end_color='2F5597', fill_type='solid') # Steel Blue
    section_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid') # Soft blue tint
    conforme_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid') # Pastel Green
    nconforme_fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid') # Pastel Rose
    
    thin_border_side = Side(border_style='thin', color='D9D9D9')
    thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    
    # Title Row
    ws.merge_cells('A1:D1')
    ws['A1'] = "RELATÓRIO DE CHECKLIST DE INÍCIO DE TURNO"
    ws['A1'].font = title_font
    ws['A1'].fill = title_fill
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 45
    
    # Helper to check timezone formatting safety
    def format_dt(dt):
        return dt.astimezone().strftime("%d/%m/%Y %H:%M:%S") if dt else "-"

    # Metadata rows
    metadata = [
        ("ID da Sessão:", session.id, "Máquina:", session.machine.name),
        ("Inspetor:", f"{session.inspector.username} ({session.inspector.specialty})", "Nome do Líder:", session.leader.get_full_name() or session.leader.username if session.leader else "-"),
        ("Status Atual:", session.get_status_display(), "Criado em:", format_dt(session.created_at)),
        ("Iniciado em:", format_dt(session.started_at), "Finalizado em:", format_dt(session.completed_at))
    ]
    
    current_row = 3
    for label1, val1, label2, val2 in metadata:
        ws.cell(row=current_row, column=1, value=label1).font = meta_label_font
        ws.cell(row=current_row, column=2, value=val1).font = meta_value_font
        ws.cell(row=current_row, column=3, value=label2).font = meta_label_font
        ws.cell(row=current_row, column=4, value=val2).font = meta_value_font
        
        for col in range(1, 5):
            ws.cell(row=current_row, column=col).border = thin_border
            ws.cell(row=current_row, column=col).alignment = Alignment(vertical='center')
            
        ws.row_dimensions[current_row].height = 22
        current_row += 1
        
    # Get pause reason history
    pause_logs = session.timeline_logs.filter(action='PAUSE').order_by('timestamp')
    pause_details = []
    for log in pause_logs:
        dt_str = log.timestamp.astimezone().strftime("%d/%m/%Y %H:%M")
        reason = log.pause_reason or "Sem motivo"
        pause_details.append(f"[{dt_str}] {reason}")
    pause_history_str = " | ".join(pause_details) if pause_details else "Nenhuma pausa registrada"

    ws.cell(row=current_row, column=1, value="Histórico de Pausas:").font = meta_label_font
    ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=4)
    ws.cell(row=current_row, column=2, value=pause_history_str).font = meta_value_font
    
    for col in range(1, 5):
        ws.cell(row=current_row, column=col).border = thin_border
        ws.cell(row=current_row, column=col).alignment = Alignment(vertical='center', wrap_text=True)
    ws.row_dimensions[current_row].height = None
    current_row += 1

    ws.cell(row=current_row, column=1, value="Obs. Gerais da Máquina:").font = meta_label_font
    ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=4)
    ws.cell(row=current_row, column=2, value=session.general_observations or "Nenhuma observação registrada").font = meta_value_font
    
    for col in range(1, 5):
        ws.cell(row=current_row, column=col).border = thin_border
        ws.cell(row=current_row, column=col).alignment = Alignment(vertical='center', wrap_text=True)
    ws.row_dimensions[current_row].height = None
    current_row += 1
        
    current_row += 1 # Empty row
    # Section header columns
    headers = ["Seção", "Item de Inspeção", "Status", "Observações"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center' if col_idx == 3 else 'left', vertical='center')
        cell.border = thin_border
    ws.row_dimensions[current_row].height = 25
    current_row += 1
    
    sections_display = {
        'HIDRAULICA': 'Hidráulica (Mecânico)',
        'PNEUMATICA': 'Pneumática (Mecânico)',
        'SISTEMA_VULCANIZACAO': 'Sistema de Vulcanização (Mecânico)',
        'SISTEMA_VAPOR': 'Sistema de Vapor (Mecânico)',
        'ESTRUTURA': 'Estrutura (Mecânico)',
        'ELETRICA': 'Elétrica (Eletricista)',
    }
    
    # Writing inspection rows
    for item in items:
        # Seção
        c1 = ws.cell(row=current_row, column=1, value=sections_display.get(item.section, item.section))
        c1.font = item_font
        c1.border = thin_border
        c1.alignment = Alignment(vertical='center')
        
        # Item Description
        c2 = ws.cell(row=current_row, column=2, value=item.item_name)
        c2.font = item_font
        c2.border = thin_border
        c2.alignment = Alignment(vertical='center')
        
        # Status
        status_text = "Conforme (C)" if item.status == 'C' else "Não Conforme (NC)" if item.status == 'NC' else "Não Respondido"
        c3 = ws.cell(row=current_row, column=3, value=status_text)
        c3.font = item_font
        c3.border = thin_border
        c3.alignment = Alignment(horizontal='center', vertical='center')
        if item.status == 'C':
            c3.fill = conforme_fill
        elif item.status == 'NC':
            c3.fill = nconforme_fill
            
        # Observations
        c4 = ws.cell(row=current_row, column=4, value=item.observations or "")
        c4.font = item_font
        c4.border = thin_border
        c4.alignment = Alignment(vertical='center')
        
        ws.row_dimensions[current_row].height = 22
        current_row += 1
        
    # Manual dimensions to ensure text wraps and fits well
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 55
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 45
    
    # Save Workbook to buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    safe_name = session.machine.name.replace(" ", "_").replace("/", "-")
    filename = f"Relatorio_Checklist_{safe_name}_ID_{session.id}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response

@login_required
def export_dashboard_excel(request):
    if not request.user.can_export_excel:
        return HttpResponseForbidden("Você não tem permissão para exportar o histórico para Excel.")

    # --- FILTRO POR PERÍODO (mesma lógica da dashboard) ---
    today = timezone.localdate()
    default_inicio = today.replace(day=1)
    default_fim = today

    data_inicio_str = request.GET.get('data_inicio', '')
    data_fim_str = request.GET.get('data_fim', '')

    try:
        data_inicio = datetime.strptime(data_inicio_str, '%Y-%m-%d').date() if data_inicio_str else default_inicio
    except ValueError:
        data_inicio = default_inicio

    try:
        data_fim = datetime.strptime(data_fim_str, '%Y-%m-%d').date() if data_fim_str else default_fim
    except ValueError:
        data_fim = default_fim

    dt_inicio = timezone.make_aware(datetime.combine(data_inicio, dt_time.min))
    dt_fim = timezone.make_aware(datetime.combine(data_fim, dt_time.max))

    maquina_id = request.GET.get('maquina', '')
    inspector_id = request.GET.get('inspector', '')
    leader_id = request.GET.get('leader_filter', '')
    status_filter = request.GET.get('status_filter', '')
    nc_query = request.GET.get('nc_query', '')

    sessions = ChecklistSession.objects.filter(
        created_at__gte=dt_inicio, created_at__lte=dt_fim
    ).select_related('machine', 'inspector').order_by('-created_at')

    if maquina_id:
        sessions = sessions.filter(machine_id=maquina_id)
    if inspector_id:
        sessions = sessions.filter(inspector_id=inspector_id)
    if leader_id:
        sessions = sessions.filter(leader_id=leader_id)
    if status_filter:
        sessions = sessions.filter(status=status_filter)
    if nc_query:
        sessions = sessions.filter(items__status='NC', items__item_name__icontains=nc_query).distinct()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Histórico Checklists"
    
    title_font = Font(name='Segoe UI', size=16, bold=True, color='FFFFFF')
    header_font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
    cell_font = Font(name='Segoe UI', size=10)
    
    title_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    header_fill = PatternFill(start_color='2F5597', end_color='2F5597', fill_type='solid')
    
    thin_border_side = Side(border_style='thin', color='D9D9D9')
    thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    
    # Title Row
    ws.merge_cells('A1:I1')
    ws['A1'] = "HISTÓRICO CONSOLIDADO DE CHECKLISTS"
    ws['A1'].font = title_font
    ws['A1'].fill = title_fill
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 45
    
    # Header columns
    headers = ["ID", "Data Criação", "Máquina", "Inspetor", "Líder", "Status", "Não Conformidades", "Histórico de Pausas", "Observações Gerais"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center' if col_idx in [1, 2, 6] else 'left', vertical='center')
        cell.border = thin_border
    ws.row_dimensions[3].height = 25
    
    current_row = 4
    for s in sessions:
        nc_items = s.items.filter(status='NC')
        nc_summary = ", ".join([item.item_name for item in nc_items]) if nc_items.exists() else "Nenhuma"
        
        # Concatenate pause logs
        pause_logs = s.timeline_logs.filter(action='PAUSE').order_by('timestamp')
        pause_details = []
        for log in pause_logs:
            dt_str = log.timestamp.astimezone().strftime("%d/%m/%Y %H:%M")
            reason = log.pause_reason or "Sem motivo"
            pause_details.append(f"[{dt_str}] {reason}")
        pause_history_str = " | ".join(pause_details) if pause_details else "Nenhuma"
 
        ws.cell(row=current_row, column=1, value=s.id).alignment = Alignment(horizontal='center')
        ws.cell(row=current_row, column=2, value=s.created_at.astimezone().strftime("%d/%m/%Y %H:%M")).alignment = Alignment(horizontal='center')
        ws.cell(row=current_row, column=3, value=s.machine.name)
        ws.cell(row=current_row, column=4, value=f"{s.inspector.username} ({s.inspector.specialty})")
        ws.cell(row=current_row, column=5, value=s.leader.get_full_name() or s.leader.username if s.leader else "-")
        ws.cell(row=current_row, column=6, value=s.get_status_display()).alignment = Alignment(horizontal='center')
        ws.cell(row=current_row, column=7, value=nc_summary)
        ws.cell(row=current_row, column=8, value=pause_history_str)
        ws.cell(row=current_row, column=9, value=s.general_observations or "")
        
        for col in range(1, 10):
            cell = ws.cell(row=current_row, column=col)
            cell.font = cell_font
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center')
            
        ws.row_dimensions[current_row].height = 20
        current_row += 1
        
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 22
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 60
    ws.column_dimensions['H'].width = 50
    ws.column_dimensions['I'].width = 40
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    safe_inicio = data_inicio.strftime('%d-%m-%Y')
    safe_fim = data_fim.strftime('%d-%m-%Y')
    response['Content-Disposition'] = f'attachment; filename="Historico_Checklists_{safe_inicio}_a_{safe_fim}.xlsx"'
    return response

# =====================================================================
# TIMEOUT POR INATIVIDADE - SALVAR E PAUSAR ANTES DO LOGOUT
# =====================================================================

@login_required
def timeout_pause_checklist(request):
    """
    View AJAX chamada pelo JavaScript do front-end imediatamente antes
    do redirect de logout automático do django-session-security.

    Busca o checklist IN_PROGRESS do usuário logado e o pausa com o
    motivo 'Pausa automática: Timeout por inatividade', garantindo que
    nenhum dado em andamento seja perdido.

    Não aceita session_id do cliente — busca pelo próprio request.user
    para evitar manipulação de requisição.
    """
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Método não permitido.'}, status=405)

    # Busca o checklist ativo do usuário autenticado
    active_session = ChecklistSession.objects.filter(
        inspector=request.user,
        status='IN_PROGRESS'
    ).first()

    if not active_session:
        # Nenhum checklist em andamento — logout pode prosseguir normalmente
        return JsonResponse({'status': 'ok', 'message': 'Nenhum checklist ativo. Logout liberado.'})

    MOTIVO_TIMEOUT = "Pausa automática: Timeout por inatividade"

    try:
        with transaction.atomic():
            active_session.status = 'PAUSED'
            active_session.paused_at = timezone.now()
            active_session.save()

            ChecklistTimelineLog.objects.create(
                session=active_session,
                user=request.user,
                action='PAUSE',
                pause_reason=MOTIVO_TIMEOUT
            )
        return JsonResponse({
            'status': 'ok',
            'message': f'Checklist #{active_session.id} pausado automaticamente por inatividade.',
            'session_id': active_session.id,
        })
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
